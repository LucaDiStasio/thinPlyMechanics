#!/usr/bin/env Python
# -*- coding: utf-8 -*-

'''
=====================================================================================

Copyright (c) 2016-2018 Université de Lorraine & Luleå tekniska universitet
Author: Luca Di Stasio <luca.distasio@gmail.com>
                       <luca.distasio@ingpec.eu>

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.

=====================================================================================

DESCRIPTION



Tested with Python 2.7 Anaconda 2.4.1 (64-bit) distribution in Windows 10.

'''

import sys
import os
from os.path import isfile, join, exists
#from os import listdir, stat, makedirs
from datetime import datetime
from time import strftime
#from platform import platform
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
import numpy as np
from scipy import optimize
import matplotlib.pyplot as plt
from matplotlib import rc
rc('font',**{'family':'sans-serif','sans-serif':['Helvetica']})
# for Palatino and other serif fonts use:
#rc('font',**{'family':'serif','serif':['Palatino']})
rc('text', usetex=True)


def model(x,A,B,C,D):
    return (A*np.sin(B*x+C)+D)

def linear(x,A,B):
    return (A*x+B)

def readData(wd,workbook,boundaryCase):
    wb = load_workbook(filename=join(wd,workbook), data_only=True, read_only=True)
    gIvcctWorksheet = wb['GI-VCCT']
    gIjintWorksheet = wb['GI-Jint']
    gIIvcctWorksheet = wb['GII-VCCT']
    gTOTvcctWorksheet = wb['GTOT-VCCT']
    gTOTjintWorksheet = wb['GTOT-Jint']
    czWorksheet = wb['ContactZone']
    data = {}
    Vf = [0.000079,0.0001,0.2,0.3,0.4,0.5,0.55,0.6,0.65]
    Vfcolumns = {0.000079:'F',0.0001:'G',0.2:'H',0.3:'I',0.4:'J',0.5:'K',0.55:'L',0.6:'M',0.65:'N'}
    theta = [10.0,20.0,30.0,40.0,50.0,60.0,70.0,80.0,90.0,100.0,110.0,120.0,130.0,140.0,150.0]
    nTheta = len(theta)
    initTheta = 4
    data['GI'] = {}
    data['GII'] = {}
    data['GTOT'] = {}
    data['CZ'] = {}
    data['GI']['VCCT'] = {}
    data['GI']['Jint'] = {}
    data['GII']['VCCT'] = {}
    data['GTOT']['VCCT'] = {}
    data['GTOT']['Jint'] = {}
    for c,case in enumerate(boundaryCase):
        data['CZ'][case] = {}
        data['GI']['VCCT'][case] = {}
        data['GI']['Jint'][case] = {}
        data['GII']['VCCT'][case] = {}
        data['GTOT']['VCCT'][case] = {}
        data['GTOT']['Jint'][case] = {}
        for v,value in enumerate(Vf):
            values = []
            for n in range(0,nTheta):
                values.append(gIvcctWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['GI']['VCCT'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
            values = []
            for n in range(0,nTheta):
                values.append(gIjintWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['GI']['Jint'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
            values = []
            for n in range(0,nTheta):
                values.append(gIIvcctWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['GII']['VCCT'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
            values = []
            for n in range(0,nTheta):
                values.append(gTOTvcctWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['GTOT']['VCCT'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
            values = []
            for n in range(0,nTheta):
                values.append(gTOTjintWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['GTOT']['Jint'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
            values = []
            for n in range(0,nTheta):
                values.append(czWorksheet[Vfcolumns[value]+str(4+c*nTheta+n)].value)
            data['CZ'][case][v+1] = {'Vf':value,'theta':theta,'values':values}
    return data

def interpolateData(outdir,data,boundaryCase):
    plt.close("all")
    for c,case in enumerate(boundaryCase):
        for vI,vfData in enumerate(data['GI']['VCCT'][case].keys()):
            print('GI VCCT, Vf ' + str(data['GI']['VCCT'][case][vfData]['Vf']) + ', ' + case)
            czStart = -1
            for a,angle in enumerate(data['CZ'][case][vfData]['values']):
                if angle>0.0:
                    czStart = a
                    break
            filename = datetime.now().strftime('%Y-%m-%d') + '_GI-VCCT-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['GI']['VCCT'][case][vfData]['Vf'])
            xs = data['GI']['VCCT'][case][vfData]['theta'][:czStart]
            ys = data['GI']['VCCT'][case][vfData]['values'][:czStart]
            maxIndex = 0
            maxValue = data['GI']['VCCT'][case][vfData]['values'][maxIndex]
            for i,y in enumerate(ys):
                if y>maxValue:
                    maxIndex = i
                    maxValue = y
            try:
                res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],0.0,0.0],method='dogbox')
                stderr = np.sqrt(np.diag(cov))
                angles = np.linspace(0.0, xs[-1]+5, num=300)
                data['GI']['VCCT'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                data['GI']['VCCT'][case][vfData]['cov'] = cov
                data['GI']['VCCT'][case][vfData]['std'] = stderr
                data['GI']['VCCT'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                plt.figure()
                plt.plot(data['GI']['VCCT'][case][vfData]['theta'], data['GI']['VCCT'][case][vfData]['values'], 'ko')
                plt.plot(angles, model(angles, *res), 'b-')
                plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                plt.title(r'Interpolation of GI-VCCT, ' + case + ', ' + str(data['GI']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                plt.legend(('data', 'interpolant'), loc='best')
                plt.grid(True)
                plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            except Exception, error:
                try:
                    res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],-0.5*np.pi,0.0],method='dogbox')
                    stderr = np.sqrt(np.diag(cov))
                    angles = np.linspace(0.0, xs[-1]+5, num=300)
                    data['GI']['VCCT'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                    data['GI']['VCCT'][case][vfData]['cov'] = cov
                    data['GI']['VCCT'][case][vfData]['std'] = stderr
                    data['GI']['VCCT'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                    plt.figure()
                    plt.plot(data['GI']['VCCT'][case][vfData]['theta'], data['GI']['VCCT'][case][vfData]['values'], 'ko')
                    plt.plot(angles, model(angles, *res), 'b-')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GI-VCCT, ' + case + ', ' + str(data['GI']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
                    except Exception, error:
                        data['GI']['VCCT'][case][vfData]['coeff'] = [0,0,0,0]
                        data['GI']['VCCT'][case][vfData]['cov'] = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
                        data['GI']['VCCT'][case][vfData]['std'] = [0,0,0,0]
                        data['GI']['VCCT'][case][vfData]['valueAtNoDebond'] = 0
                        plt.figure()
                        plt.plot(data['GI']['VCCT'][case][vfData]['theta'], data['GI']['VCCT'][case][vfData]['values'], 'ko')
                        plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                        plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                        plt.title(r'Interpolation of GI-VCCT, ' + case + ', ' + str(data['GI']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                        plt.legend(('data', 'interpolant: FAILED'), loc='best')
                        plt.grid(True)
                        plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                        sys.exc_clear()
            plt.close("all")
        for vI,vfData in enumerate(data['GI']['Jint'][case].keys()):
            print('GI Jint, Vf ' + str(data['GI']['Jint'][case][vfData]['Vf']) + ', ' + case)
            czStart = -1
            for a,angle in enumerate(data['CZ'][case][vfData]['values']):
                if angle>0.0:
                    czStart = a
                    break
            filename = datetime.now().strftime('%Y-%m-%d') + '_GI-Jint-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['GI']['Jint'][case][vfData]['Vf'])
            xs = data['GI']['Jint'][case][vfData]['theta'][:czStart]
            ys = data['GI']['Jint'][case][vfData]['values'][:czStart]
            maxIndex = 0
            maxValue = data['GI']['Jint'][case][vfData]['values'][maxIndex]
            for i,y in enumerate(ys):
                if y>maxValue:
                    maxIndex = i
                    maxValue = y
            try:
                res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],0.0,0.0],method='dogbox')
                stderr = np.sqrt(np.diag(cov))
                angles = np.linspace(0.0, xs[-1]+5, num=300)
                data['GI']['Jint'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                data['GI']['Jint'][case][vfData]['cov'] = cov
                data['GI']['Jint'][case][vfData]['std'] = stderr
                data['GI']['Jint'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                plt.figure()
                plt.plot(data['GI']['Jint'][case][vfData]['theta'], data['GI']['Jint'][case][vfData]['values'], 'ko')
                plt.plot(angles, model(angles, *res), 'b-')
                plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                plt.title(r'Interpolation of GI-Jint, ' + case + ', ' + str(data['GI']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                plt.legend(('data', 'interpolant'), loc='best')
                plt.grid(True)
                plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            except Exception, error:
                try:
                    res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],-0.5*np.pi,0.0],method='dogbox')
                    stderr = np.sqrt(np.diag(cov))
                    angles = np.linspace(0.0, xs[-1]+5, num=300)
                    data['GI']['Jint'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                    data['GI']['Jint'][case][vfData]['cov'] = cov
                    data['GI']['Jint'][case][vfData]['std'] = stderr
                    data['GI']['Jint'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                    plt.figure()
                    plt.plot(data['GI']['Jint'][case][vfData]['theta'], data['GI']['Jint'][case][vfData]['values'], 'ko')
                    plt.plot(angles, model(angles, *res), 'b-')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GI-Jint, ' + case + ', ' + str(data['GI']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
                except Exception, error:
                    data['GI']['Jint'][case][vfData]['coeff'] = [0,0,0,0]
                    data['GI']['Jint'][case][vfData]['cov'] = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
                    data['GI']['Jint'][case][vfData]['std'] = [0,0,0,0]
                    data['GI']['Jint'][case][vfData]['valueAtNoDebond'] = 0
                    plt.figure()
                    plt.plot(data['GI']['Jint'][case][vfData]['theta'], data['GI']['Jint'][case][vfData]['values'], 'ko')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{I} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GI-Jint, ' + case + ', ' + str(data['GI']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant:FAILED'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
            plt.close("all")
        for vI,vfData in enumerate(data['GII']['VCCT'][case].keys()):
            print('GII, Vf ' + str(data['GII']['VCCT'][case][vfData]['Vf']) + ', ' + case)
            filename = datetime.now().strftime('%Y-%m-%d') + '_GII-VCCT-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['GII']['VCCT'][case][vfData]['Vf'])
            refValue = data['GII']['VCCT'][case][vfData]['values'][0]
            rangeEnd = -1
            for v,value in enumerate(data['GII']['VCCT'][case][vfData]['values']):
                if value<0.95*refValue:
                    rangeEnd = v
                    break
            xs = data['GII']['VCCT'][case][vfData]['theta'][:rangeEnd]
            ys = data['GII']['VCCT'][case][vfData]['values'][:rangeEnd]
            maxIndex = 0
            maxValue = ys[maxIndex]
            for i,y in enumerate(ys):
                if y>maxValue:
                    maxIndex = i
                    maxValue = y
            res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],0.0,0.0],method='dogbox')
            stderr = np.sqrt(np.diag(cov))
            angles = np.linspace(0.0, xs[-1]+5, num=300)
            data['GII']['VCCT'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
            data['GII']['VCCT'][case][vfData]['cov'] = cov
            data['GII']['VCCT'][case][vfData]['std'] = stderr
            data['GII']['VCCT'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
            plt.figure()
            plt.plot(data['GII']['VCCT'][case][vfData]['theta'], data['GII']['VCCT'][case][vfData]['values'], 'ko')
            plt.plot(angles, model(angles, *res), 'b-')
            plt.xlabel(r'$\Delta\theta [^{\circ}]$')
            plt.ylabel(r'$G_{II} [\frac{J}{m^{2}}]$')
            plt.title(r'Interpolation of GII-VCCT, ' + case + ', ' + str(data['GII']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
            plt.legend(('data', 'interpolant'), loc='best')
            plt.grid(True)
            plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            plt.close("all")
        for vI,vfData in enumerate(data['GTOT']['VCCT'][case].keys()):
            print('GTOT VCCT, Vf ' + str(data['GTOT']['VCCT'][case][vfData]['Vf']) + ', ' + case)
            filename = datetime.now().strftime('%Y-%m-%d') + '_GTOT-VCCT-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['GTOT']['VCCT'][case][vfData]['Vf'])
            refValue = data['GII']['VCCT'][case][vfData]['values'][0]
            rangeEnd = -1
            for v,value in enumerate(data['GII']['VCCT'][case][vfData]['values']):
                if value<0.95*refValue:
                    rangeEnd = v
                    break
            xs = data['GTOT']['VCCT'][case][vfData]['theta'][:rangeEnd]
            ys = data['GTOT']['VCCT'][case][vfData]['values'][:rangeEnd]
            maxIndex = 0
            maxValue = ys[maxIndex]
            for i,y in enumerate(ys):
                if y>maxValue:
                    maxIndex = i
                    maxValue = y
            try:
                res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],0.0,0.0],method='dogbox')
                stderr = np.sqrt(np.diag(cov))
                angles = np.linspace(0.0, xs[-1]+5, num=300)
                data['GTOT']['VCCT'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                data['GTOT']['VCCT'][case][vfData]['cov'] = cov
                data['GTOT']['VCCT'][case][vfData]['std'] = stderr
                data['GTOT']['VCCT'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                plt.figure()
                plt.plot(data['GTOT']['VCCT'][case][vfData]['theta'], data['GTOT']['VCCT'][case][vfData]['values'], 'ko')
                plt.plot(angles, model(angles, *res), 'b-')
                plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                plt.title(r'Interpolation of GTOT-VCCT, ' + case + ', ' + str(data['GTOT']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                plt.legend(('data', 'interpolant'), loc='best')
                plt.grid(True)
                plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            except Exception, error:
                try:
                    res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],-0.5*np.pi,0.0],method='dogbox')
                    stderr = np.sqrt(np.diag(cov))
                    angles = np.linspace(0.0, xs[-1]+5, num=300)
                    data['GTOT']['VCCT'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                    data['GTOT']['VCCT'][case][vfData]['cov'] = cov
                    data['GTOT']['VCCT'][case][vfData]['std'] = stderr
                    data['GTOT']['VCCT'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                    plt.figure()
                    plt.plot(data['GTOT']['VCCT'][case][vfData]['theta'], data['GTOT']['VCCT'][case][vfData]['values'], 'ko')
                    plt.plot(angles, model(angles, *res), 'b-')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GTOT-VCCT, ' + case + ', ' + str(data['GTOT']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
                except Exception, error:
                    data['GTOT']['VCCT'][case][vfData]['coeff'] = [0,0,0,0]
                    data['GTOT']['VCCT'][case][vfData]['cov'] = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
                    data['GTOT']['VCCT'][case][vfData]['std'] = [0,0,0,0]
                    data['GTOT']['VCCT'][case][vfData]['valueAtNoDebond'] = 0
                    plt.figure()
                    plt.plot(data['GTOT']['VCCT'][case][vfData]['theta'], data['GTOT']['VCCT'][case][vfData]['values'], 'ko')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GTOT-VCCT, ' + case + ', ' + str(data['GTOT']['VCCT'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant: FAILED'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
            plt.close("all")
        for vI,vfData in enumerate(data['GTOT']['Jint'][case].keys()):
            print('GTOT Jint, Vf ' + str(data['GTOT']['Jint'][case][vfData]['Vf']) + ', ' + case)
            filename = datetime.now().strftime('%Y-%m-%d') + '_GTOT-Jint-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['GTOT']['Jint'][case][vfData]['Vf'])
            refValue = data['GII']['VCCT'][case][vfData]['values'][0]
            rangeEnd = -1
            for v,value in enumerate(data['GII']['VCCT'][case][vfData]['values']):
                if value<0.95*refValue:
                    rangeEnd = v
                    break
            xs = data['GTOT']['Jint'][case][vfData]['theta'][:rangeEnd]
            ys = data['GTOT']['Jint'][case][vfData]['values'][:rangeEnd]
            maxIndex = 0
            maxValue = ys[maxIndex]
            for i,y in enumerate(ys):
                if y>maxValue:
                    maxIndex = i
                    maxValue = y
            try:
                res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],0.0,0.0],method='dogbox')
                stderr = np.sqrt(np.diag(cov))
                angles = np.linspace(0.0, xs[-1]+5, num=300)
                data['GTOT']['Jint'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                data['GTOT']['Jint'][case][vfData]['cov'] = cov
                data['GTOT']['Jint'][case][vfData]['std'] = stderr
                data['GTOT']['Jint'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                plt.figure()
                plt.plot(data['GTOT']['Jint'][case][vfData]['theta'], data['GTOT']['Jint'][case][vfData]['values'], 'ko')
                plt.plot(angles, model(angles, *res), 'b-')
                plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                plt.title(r'Interpolation of GTOT-Jint, ' + case + ', ' + str(data['GTOT']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                plt.legend(('data', 'interpolant'), loc='best')
                plt.grid(True)
                plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            except Exception, error:
                try:
                    res, cov = optimize.curve_fit(model,xs,ys,p0=[maxValue,1.0/xs[maxIndex],-0.5*np.pi,0.0],method='dogbox')
                    stderr = np.sqrt(np.diag(cov))
                    angles = np.linspace(0.0, xs[-1]+5, num=300)
                    data['GTOT']['Jint'][case][vfData]['coeff'] = [res[0],res[1],res[2]*180.0/np.pi,res[-1]]
                    data['GTOT']['Jint'][case][vfData]['cov'] = cov
                    data['GTOT']['Jint'][case][vfData]['std'] = stderr
                    data['GTOT']['Jint'][case][vfData]['valueAtNoDebond'] = model(0.0, *res)
                    plt.figure()
                    plt.plot(data['GTOT']['Jint'][case][vfData]['theta'], data['GTOT']['Jint'][case][vfData]['values'], 'ko')
                    plt.plot(angles, model(angles, *res), 'b-')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GTOT-Jint, ' + case + ', ' + str(data['GTOT']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
                except Exception, error:
                    data['GTOT']['Jint'][case][vfData]['coeff'] = [0,0,0,0]
                    data['GTOT']['Jint'][case][vfData]['cov'] = [[0,0,0,0],[0,0,0,0],[0,0,0,0],[0,0,0,0]]
                    data['GTOT']['Jint'][case][vfData]['std'] = [0,0,0,0]
                    data['GTOT']['Jint'][case][vfData]['valueAtNoDebond'] = 0
                    plt.figure()
                    plt.plot(data['GTOT']['Jint'][case][vfData]['theta'], data['GTOT']['Jint'][case][vfData]['values'], 'ko')
                    plt.xlabel(r'$\Delta\theta [^{\circ}]$')
                    plt.ylabel(r'$G_{TOT} [\frac{J}{m^{2}}]$')
                    plt.title(r'Interpolation of GTOT-Jint, ' + case + ', ' + str(data['GTOT']['Jint'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
                    plt.legend(('data', 'interpolant: FAILED'), loc='best')
                    plt.grid(True)
                    plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
                    sys.exc_clear()
            plt.close("all")
        for vI,vfData in enumerate(data['CZ'][case].keys()):
            print('CZ, Vf ' + str(data['CZ'][case][vfData]['Vf']) + ', ' + case)
            czStart = -1
            for a,angle in enumerate(data['CZ'][case][vfData]['values']):
                if angle>0.0:
                    czStart = a
                    break
            filename = datetime.now().strftime('%Y-%m-%d') + '_ContactZone-Interpolation_' + case + '_' + str(vfData) + '_Vf' + str(data['CZ'][case][vfData]['Vf'])
            xs = data['CZ'][case][vfData]['theta'][czStart:]
            phis = []
            for p,phi in enumerate(data['CZ'][case][vfData]['values']):
                phis.append(100*phi/data['CZ'][case][vfData]['theta'][p])
            ys = phis[czStart:]
            res, cov = optimize.curve_fit(linear,xs,ys,method='dogbox')
            stderr = np.sqrt(np.diag(cov))
            angles = np.linspace(xs[0], xs[-1]+5, num=300)
            data['CZ'][case][vfData]['coeff'] = res
            data['CZ'][case][vfData]['cov'] = cov
            data['CZ'][case][vfData]['std'] = stderr
            plt.figure()
            plt.plot(data['CZ'][case][vfData]['theta'], phis, 'ko')
            plt.plot(angles, linear(angles, *res), 'b-')
            plt.xlabel(r'$\Delta\theta [^{\circ}]$')
            plt.ylabel(r'$\frac{\Delta\Phi}{\Delta\theta} [\%]$')
            plt.title(r'Interpolation of normalized contact zone size, ' + case + ', ' + str(data['CZ'][case][vfData]['Vf']*100.0) + '\% $V_{f}$')
            plt.legend(('data', 'interpolant'), loc='best')
            plt.grid(True)
            plt.savefig(join(outdir,filename + '.png'), bbox_inches='tight')
            plt.close("all")
    return data

def writeData(outdir,workbookname,data,boundaryCase):
    workbook = xlsxwriter.Workbook(join(outdir,workbookname))
    gIvcctWorksheet = workbook.add_worksheet('GI-VCCT')
    gIjintWorksheet = workbook.add_worksheet('GI-Jint')
    gIIvcctWorksheet = workbook.add_worksheet('GII-VCCT')
    gTOTvcctWorksheet = workbook.add_worksheet('GTOT-VCCT')
    gTOTjintWorksheet = workbook.add_worksheet('GTOT-Jint')
    wsNames = ['GI-VCCT','GI-Jint','GII-VCCT','GTOT-VCCT','GTOT-Jint']
    gIvcctWorksheet.write(0,0,'GI-VCCT')
    gIjintWorksheet.write(0,0,'GI-Jint')
    gIIvcctWorksheet.write(0,0,'GII-VCCT')
    gTOTvcctWorksheet.write(0,0,'GTOT-VCCT')
    gTOTjintWorksheet.write(0,0,'GTOT-Jint')    
    for ws in [gIvcctWorksheet,gIjintWorksheet,gIIvcctWorksheet,gTOTvcctWorksheet,gTOTjintWorksheet]:
        ws.write(1,0,'BOUNDARY CONDITIONS')
        ws.write(2,0,'NORTH')
        ws.write(2,1,'SOUTH')
        ws.write(2,2,'EAST')
        ws.write(2,3,'WEST')
        ws.write(2,4,'Vf [%]')
        ws.write(2,5,'Value at no debond [J/m^2]')
        ws.write(1,6,'Coefficients')
        ws.write(2,6,'A [J/m^2]')
        ws.write(2,7,'B [1/deg]')
        ws.write(2,8,'C [deg]')
        ws.write(2,9,'D [J/m^2]')
        ws.write(1,10,'Covariance')
        ws.write(2,10,'cov(1,1)')
        ws.write(2,11,'cov(1,2)')
        ws.write(2,12,'cov(1,3)')
        ws.write(2,13,'cov(1,4)')
        ws.write(2,14,'cov(2,1)')
        ws.write(2,15,'cov(2,2)')
        ws.write(2,16,'cov(2,3)')
        ws.write(2,17,'cov(2,4)')
        ws.write(2,18,'cov(3,1)')
        ws.write(2,19,'cov(3,2)')
        ws.write(2,20,'cov(3,3)')
        ws.write(2,21,'cov(3,4)')
        ws.write(2,22,'cov(4,1)')
        ws.write(2,23,'cov(4,2)')
        ws.write(2,24,'cov(4,3)')
        ws.write(2,25,'cov(4,4)')
        ws.write(1,26,'std err')
        ws.write(2,26,'A [J/m^2]')
        ws.write(2,27,'B [1/deg]')
        ws.write(2,28,'C [deg]')
        ws.write(2,29,'D [J/m^2]')
    czWorksheet = workbook.add_worksheet('ContactZone')
    czWorksheet.write(0,0,'Contact zone')
    czWorksheet.write(0,0,'BOUNDARY CONDITIONS')
    czWorksheet.write(0,0,'NORTH')
    czWorksheet.write(0,1,'SOUTH')
    czWorksheet.write(0,2,'EAST')
    czWorksheet.write(0,3,'WEST')
    czWorksheet.write(0,4,'Vf [%]')
    czWorksheet.write(0,5,'Coefficients')
    czWorksheet.write(0,5,'A [%/deg]')
    czWorksheet.write(0,6,'B [%]')
    czWorksheet.write(0,7,'Covariance')
    czWorksheet.write(0,7,'cov(1,1)')
    czWorksheet.write(0,8,'cov(1,2)')
    czWorksheet.write(0,9,'cov(2,1)')
    czWorksheet.write(0,10,'cov(2,2)')
    czWorksheet.write(0,11,'std err')
    czWorksheet.write(0,11,'A [%/deg]')
    czWorksheet.write(0,12,'B [%]')
    Vf = [0.000079,0.0001,0.2,0.3,0.4,0.5,0.55,0.6,0.65]
    nVf = len(Vf)
    initVf = 3
    for c,case in enumerate(boundaryCase):
        for ws in [gIvcctWorksheet,gIjintWorksheet,gIIvcctWorksheet,gTOTvcctWorksheet,gTOTjintWorksheet,czWorksheet]:
            ws.write(initVf+c*nVf,0,case)
            ws.write(initVf+c*nVf,1,'Ysymm')
            ws.write(initVf+c*nVf,2,'epsx=1%')
            ws.write(initVf+c*nVf,3,'epsx=1%')
        for v,value in enumerate(Vf):
            gIvcctWorksheet.write(initVf+c*nVf+v,4,100*value)
            gIvcctWorksheet.write(initVf+c*nVf+v,5,data['GI']['VCCT'][case][v+1]['valueAtNoDebond'])
            for cIndex,coeff in enumerate(data['GI']['VCCT'][case][v+1]['coeff']):
                gIvcctWorksheet.write(initVf+c*nVf+v,6+cIndex,coeff)
            for r,row in enumerate(data['GI']['VCCT'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    gIvcctWorksheet.write(initVf+c*nVf+v,10+4*r+col,cov)
            for cIndex,coeff in enumerate(data['GI']['VCCT'][case][v+1]['std']):
                gIvcctWorksheet.write(initVf+c*nVf+v,26+cIndex,coeff)
            gIjintWorksheet.write(initVf+c*nVf+v,4,100*value)
            gIjintWorksheet.write(initVf+c*nVf+v,5,data['GI']['Jint'][case][v+1]['valueAtNoDebond'])
            for cIndex,coeff in enumerate(data['GI']['Jint'][case][v+1]['coeff']):
                gIjintWorksheet.write(initVf+c*nVf+v,6+cIndex,coeff)
            for r,row in enumerate(data['GI']['Jint'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    gIjintWorksheet.write(initVf+c*nVf+v,10+4*r+col,cov)
            for cIndex,coeff in enumerate(data['GI']['Jint'][case][v+1]['std']):
                gIjintWorksheet.write(initVf+c*nVf+v,26+cIndex,coeff)
            gIIvcctWorksheet.write(initVf+c*nVf+v,4,100*value)
            gIIvcctWorksheet.write(initVf+c*nVf+v,5,data['GII']['VCCT'][case][v+1]['valueAtNoDebond'])
            for cIndex,coeff in enumerate(data['GII']['VCCT'][case][v+1]['coeff']):
                gIIvcctWorksheet.write(initVf+c*nVf+v,6+cIndex,coeff)
            for r,row in enumerate(data['GII']['VCCT'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    gIIvcctWorksheet.write(initVf+c*nVf+v,10+4*r+col,cov)
            for cIndex,coeff in enumerate(data['GII']['VCCT'][case][v+1]['std']):
                gIIvcctWorksheet.write(initVf+c*nVf+v,26+cIndex,coeff)    
            gTOTvcctWorksheet.write(initVf+c*nVf+v,4,100*value)
            gTOTvcctWorksheet.write(initVf+c*nVf+v,5,data['GTOT']['VCCT'][case][v+1]['valueAtNoDebond'])
            for cIndex,coeff in enumerate(data['GTOT']['VCCT'][case][v+1]['coeff']):
                gTOTvcctWorksheet.write(initVf+c*nVf+v,6+cIndex,coeff)
            for r,row in enumerate(data['GTOT']['VCCT'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    gTOTvcctWorksheet.write(initVf+c*nVf+v,10+4*r+col,cov)
            for cIndex,coeff in enumerate(data['GTOT']['VCCT'][case][v+1]['std']):
                gTOTvcctWorksheet.write(initVf+c*nVf+v,26+cIndex,coeff)
            gTOTjintWorksheet.write(initVf+c*nVf+v,4,100*value)
            gTOTjintWorksheet.write(initVf+c*nVf+v,5,data['GTOT']['Jint'][case][v+1]['valueAtNoDebond'])
            for cIndex,coeff in enumerate(data['GTOT']['Jint'][case][v+1]['coeff']):
                gTOTjintWorksheet.write(initVf+c*nVf+v,6+cIndex,coeff)
            for r,row in enumerate(data['GTOT']['Jint'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    gTOTjintWorksheet.write(initVf+c*nVf+v,10+4*r+col,cov)
            for cIndex,coeff in enumerate(data['GTOT']['Jint'][case][v+1]['std']):
                gTOTjintWorksheet.write(initVf+c*nVf+v,26+cIndex,coeff)
            czWorksheet.write(initVf+c*nVf+v,4,100*value)
            for cIndex,coeff in enumerate(data['CZ'][case][v+1]['coeff']):
                czWorksheet.write(initVf+c*nVf+v,5+cIndex,coeff)
            for r,row in enumerate(data['CZ'][case][v+1]['cov']):
                for col,cov in enumerate(row):
                    czWorksheet.write(initVf+c*nVf+v,7+4*r+col,cov)
            for cIndex,coeff in enumerate(data['CZ'][case][v+1]['std']):
                czWorksheet.write(initVf+c*nVf+v,11+cIndex,coeff)
    for wsIndex,ws in enumerate([gIvcctWorksheet,gIjintWorksheet,gIIvcctWorksheet,gTOTvcctWorksheet,gTOTjintWorksheet]):
        chartDeb = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
        chartDeb.set_title ({'name': 'Value at no debond'})
        chartDeb.set_y_axis({'name': 'G(0) [J/m^2]'})
        chartDeb.set_x_axis({'name': 'Vf [%]'})
        chartA = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
        chartA.set_title ({'name': 'Amplitude'})
        chartA.set_y_axis({'name': 'A [J/m^2]'})
        chartA.set_x_axis({'name': 'Vf [%]'})
        chartB = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
        chartB.set_title ({'name': 'Spatial frequency'})
        chartB.set_y_axis({'name': 'B [1/deg]'})
        chartB.set_x_axis({'name': 'Vf [%]'})
        chartC = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
        chartB.set_title ({'name': 'Phase'})
        chartB.set_y_axis({'name': 'C [deg]'})
        chartB.set_x_axis({'name': 'Vf [%]'})
        chartD = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
        chartD.set_title ({'name': 'Offset'})
        chartD.set_y_axis({'name': 'D [J/m^2]'})
        chartD.set_x_axis({'name': 'Vf [%]'})
        for c,case in enumerate(boundaryCase):
            chartDeb.add_series({
                                'name':       case,
                                'categories': [wsNames[wsIndex],initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                                'values':     [wsNames[wsIndex],initVf+c*nVf,5,initVf+(c+1)*nVf-1,5],
                            })
            chartA.add_series({
                                'name':       case,
                                'categories': [wsNames[wsIndex],initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                                'values':     [wsNames[wsIndex],initVf+c*nVf,6,initVf+(c+1)*nVf-1,6],
                            })
            chartB.add_series({
                                'name':       case,
                                'categories': [wsNames[wsIndex],initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                                'values':     [wsNames[wsIndex],initVf+c*nVf,7,initVf+(c+1)*nVf-1,7],
                            })
            chartC.add_series({
                                'name':       case,
                                'categories': [wsNames[wsIndex],initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                                'values':     [wsNames[wsIndex],initVf+c*nVf,8,initVf+(c+1)*nVf-1,8],
                            })
            chartD.add_series({
                                'name':       case,
                                'categories': [wsNames[wsIndex],initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                                'values':     [wsNames[wsIndex],initVf+c*nVf,9,initVf+(c+1)*nVf-1,9],
                            })
        ws.insert_chart(2,32, chartDeb)
        ws.insert_chart(20,32, chartA)
        ws.insert_chart(40,32, chartB)
        ws.insert_chart(60,32, chartC)
        ws.insert_chart(80,32, chartD)
    chartA = workbook.add_chart({'type': 'scatter','subtype': 'line_with_markers'})
    chartA.set_title ({'name': 'Percentual increment of contact zone'})
    chartA.set_y_axis({'name': 'CZ/DS/unit increment [%/deg]'})
    chartA.set_x_axis({'name': 'Vf []'})
    for c,case in enumerate(boundaryCase):
        chartA.add_series({
                            'name':       case,
                            'categories': ['ContactZone',initVf+c*nVf,4,initVf+(c+1)*nVf-1,4],
                            'values':     ['ContactZone',initVf+c*nVf,5,initVf+(c+1)*nVf-1,5],
                        })
    czWorksheet.insert_chart(2,15,chartA)
    workbook.close()

def main(argv):

    inpDir = 'C:/Abaqus_WD'
    outDir = 'C:/Users/lucad/OneDrive/01_Luca/07_DocMASE/07_Data/03_FEM/Interpolation'

    inpWorkbook = 'sweepDeltathetaVffBCs-GF.xlsx'
    outWorkbook = 'sweepDeltathetaVffBCs-GF_Interpolation.xlsx'

    if not os.path.exists(outDir):
            os.mkdir(outDir)

    boundaryCases = ['free','geomcoupling','fixedv','fixedvlinearu']
    
    writeData(outDir,outWorkbook,interpolateData(outDir,readData(inpDir,inpWorkbook,boundaryCases),boundaryCases),boundaryCases)



if __name__ == "__main__":
    main(sys.argv[1:])
